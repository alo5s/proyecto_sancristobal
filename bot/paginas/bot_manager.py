"""
bot_manager.py — Automatización de cobranza y manejo de popups
"""

import time


def close_tour_popup(page):
    """
    Detects and closes the tour popup if present.
    Single quick check - no retry loops.
    """
    print("Checking for tour popup...")

    try:
        # Verify we're on the expected page
        current_url = page.url
        if "/gestion-de-cobranzas" not in current_url and "/inicio" not in current_url:
            print(f"Not on expected page, current URL: {current_url}")
            return False

        # Quick check: main page + iframes
        frames_to_check = [page] + page.frames

        for frame in frames_to_check:
            # Try exact text "Cerrar" first (most reliable)
            try:
                close_btn = frame.get_by_text("Cerrar", exact=True)
                if close_btn.is_visible(timeout=2000):
                    close_btn.click()
                    print("✓ Popup closed by exact text 'Cerrar'")
                    time.sleep(0.3)
                    return True
            except Exception:
                pass

            # Try common selectors
            selectors = [
                "button.skip-button",
                "button.skip-button.link-button",
                ".introjs-skipbutton",
                ".shepherd-cancel-icon",
                "button:has-text('Cerrar')",
                "[aria-label='Close']",
                ".modal-close",
                ".close-btn",
            ]

            for selector in selectors:
                try:
                    if frame.locator(selector).is_visible(timeout=1500):
                        frame.locator(selector).click()
                        print(f"✓ Popup closed with selector: {selector}")
                        time.sleep(0.3)
                        return True
                except Exception:
                    continue

        # Fallback: Press Escape
        try:
            page.keyboard.press("Escape")
            print("✓ Escape key pressed to close popups")
            time.sleep(0.3)
        except Exception:
            pass

        print("Tour popup not detected")
        return True  # No popup = success

    except Exception as e:
        print(f"Error checking popup: {e}")
        return False

        # Paso 1: Esperar a que la página esté estable
        try:
            page.wait_for_load_state("networkidle", timeout=10000)
        except Exception:
            pass  # Si no hay red inactiva, continuar

        # Paso 2: Reintentos para capturar popups rápidos
        max_intentos = 3
        for intento in range(max_intentos):
            print(f"Intento {intento + 1} de {max_intentos}...")

            # Recolectar todas las áreas donde buscar: página principal + iframes
            frames_to_check = [page] + page.frames

            for frame in frames_to_check:
                # Selectores comunes de botones "Cerrar" en tours y modales
                selectors = [
                    "button.skip-button",
                    "button.skip-button.link-button",
                    "xpath=//button[contains(@class, 'skip')]",
                    ".introjs-skipbutton",  # Intro.js
                    ".shepherd-cancel-icon",  # Shepherd.js
                    "button:has-text('Cerrar')",
                    "button:has-text('Saltar')",
                    "button:has-text('Skip')",
                    ".tour-close",
                    "[aria-label='Close']",
                    ".modal-close",
                    ".close-btn",
                    "button[aria-label='Cerrar']",
                ]

                # Buscar por selectores CSS/XPath
                for selector in selectors:
                    try:
                        if frame.locator(selector).is_visible(timeout=2000):
                            frame.locator(selector).click()
                            print(f"✓ Popup cerrado con selector: {selector}")
                            time.sleep(0.5)  # Esperar animación
                            return True
                    except Exception:
                        continue

                # Buscar por texto exacto "Cerrar"
                try:
                    close_btn = frame.get_by_text("Cerrar", exact=True)
                    if close_btn.is_visible(timeout=2000):
                        close_btn.click()
                        print("✓ Popup cerrado por texto exacto 'Cerrar'")
                        time.sleep(0.5)
                        return True
                except Exception:
                    pass

                # Buscar por texto parcial "Cerrar" (último recurso)
                try:
                    close_btn = frame.get_by_text("Cerrar", exact=False)
                    if close_btn.first.is_visible(timeout=2000):
                        close_btn.first.click()
                        print("✓ Popup cerrado por texto parcial 'Cerrar'")
                        time.sleep(0.5)
                        return True
                except Exception:
                    pass

            # Si no se encontró en este intento, esperar antes del siguiente
            if intento < max_intentos - 1:
                time.sleep(0.5)

        # Paso 3: Fallback - Presionar Escape para cerrar cualquier modal
        try:
            page.keyboard.press("Escape")
            print("✓ Tecla Escape presionada para cerrar popups")
            time.sleep(0.5)
        except Exception:
            pass

        print("Popup de tour no detectado")
        return False

    except Exception as e:
        print(f"Error verificando popup: {e}")
        return False


class ManagerBot:
    """Maneja las etapas de automatización (basado en Automatizacion)"""

    def __init__(self, page):
        self.page = page
        self.url_gestion = "https://productores.sancristobal.com.ar/deudas-y-cobranzas/gestion-de-cobranzas"

    def detener(self):
        """Detiene la automatización"""
        print("Automatización detenida")

    def navegar_a_inicio(self):
        """Navega a la página de inicio del portal"""
        print("Navegando a inicio...")
        self.page.goto("https://productores.sancristobal.com.ar/inicio")
        self.page.wait_for_load_state("domcontentloaded", timeout=15000)

    def selector_fecha_vencimiento_7dias(self):
        """Selecciona 'Próximos 7 días' en el dropdown de Vencimiento"""
        print("Seleccionando Próximos 7 días...")
        try:
            # Abrir dropdown de fecha
            self.page.locator("#dateRange__input").click()
            
            # Esperar y hacer clic en la opción (usar :visible para evitar ambigüedad)
            item = self.page.locator(".sc-select__item:visible", has_text="Próximos 7 días").first
            item.wait_for(state="visible", timeout=5000)
            item.click(force=True)
            
            print("✓ Próximos 7 días seleccionado")
            self.page.wait_for_timeout(1000)
            return True
        except Exception as e:
            print(f"Error seleccionando 7 días: {e}")
            return False

    def fecha_vencimiento_8dias(self):
        """Selecciona 'Próximos 8 a 15 días' en el dropdown de Vencimiento"""
        print("Seleccionando Próximos 8 a 15 días...")
        try:
            # Abrir dropdown de fecha
            self.page.locator("#dateRange__input").click()
            
            # Esperar y hacer clic en la opción (usar :visible para evitar ambigüedad)
            item = self.page.locator(".sc-select__item:visible", has_text="Próximos 8 a 15 días").first
            item.wait_for(state="visible", timeout=5000)
            item.click(force=True)
            
            print("✓ Próximos 8 a 15 días seleccionado")
            self.page.wait_for_timeout(1000)
            return True
        except Exception as e:
            print(f"Error seleccionando 8 a 15 días: {e}")
            return False

    def etapa_1(self):
        """Navega a gestión de cobranzas y aplica filtros"""
        try:
            # Navegar a gestión de cobranzas
            print("Navegando a gestión de cobranzas...")
            self.page.goto(self.url_gestion)
            self.page.wait_for_load_state("domcontentloaded", timeout=15000)
            print(f"✓ URL actual: {self.page.url}")

            # Cerrar popup por si aparece después de navegar
            close_tour_popup(self.page)

            # Aplicar filtros
            self._aplicar_filtros()
            return True
        except Exception as e:
            print(f"Error en etapa 1: {e}")
            return False

    def _aplicar_filtros(self):
        """Hace clic en Filtros, selecciona las modalidades configuradas y aplica"""
        from config_manager import get_settings
        modalidades = get_settings().value("modalidades_pago", [], type=list)

        if not modalidades:
            print("Aviso: No hay modalidades de pago configuradas en config.json (modalidades_pago)")
            return

        try:
            print(f"Abriendo filtros para: {', '.join(modalidades)}...")
            filtros_btn = self.page.locator("#filtros-gestion-cobranzas").first
            filtros_btn.wait_for(state="visible", timeout=10000)
            filtros_btn.click()
            print("✓ Botón Filtros clickeado")

            self.page.wait_for_timeout(1500)

            print("Buscando input de modalidad de pago...")
            input_filtro = self.page.locator("input[formcontrolname='paymentMethods']").first
            if not input_filtro.is_visible(timeout=2000):
                input_filtro = self.page.locator("ng-select[formcontrolname='paymentMethods'] input").first

            for i, modalidad in enumerate(modalidades):
                if i > 0:
                    self.page.wait_for_timeout(500)

                input_filtro.click()
                input_filtro.fill("")
                input_filtro.fill(modalidad)
                print(f"Escribiendo {modalidad}...")

                self.page.wait_for_timeout(1500)

                option = self.page.locator("ng-dropdown-panel .ng-option").filter(has_text=modalidad).first
                option.wait_for(state="visible", timeout=5000)
                option.click()
                print(f"✓ {modalidad} seleccionado")

            self.page.wait_for_timeout(800)
            aplicar_btn = self.page.locator("#aplicar-filtros-gestion-cobranzas").first
            aplicar_btn.wait_for(state="visible", timeout=5000)
            aplicar_btn.click()
            print("✓ Filtros aplicados")

            self.page.wait_for_timeout(2000)

            print("Configurando paginación a 100...")
            try:
                select_container = self.page.locator("#pageItem__input").locator("..")
                select_container.click(force=True)
                dropdown_items = self.page.locator(".sc-select__item:visible")
                dropdown_items.first.wait_for(timeout=5000)
                self.page.locator(".sc-select__item:visible", has_text="100").first.click(force=True)
                print("✓ Paginación configurada a 100 items")
            except Exception as e:
                print(f"Aviso: No se pudo configurar paginación: {e}")

            try:
                filtros_panel = self.page.locator("#filtros-gestion-cobranzas").first
                filtros_panel.wait_for(state="hidden", timeout=5000)
                print("✓ Panel de filtros cerrado")
            except Exception:
                print("Aviso: No se pudo verificar cierre de panel")

        except Exception as e:
            print(f"Error aplicando filtros: {e}")
            try:
                self.page.screenshot(path="debug_filtros.png")
                print("Screenshot guardado: debug_filtros.png")
            except:
                pass


    def etapa_2(self) -> bool:
        """
        Procesa clientes según configuración guardada.
        Lee 'venc_7dias' y 'venc_8dias' de config.json
        """
        from config_manager import get_settings
        s = get_settings()
        dias_7 = s.value("venc_7dias", True, type=bool)
        dias_8 = s.value("venc_8dias", False, type=bool)
        
        print(f"Iniciando etapa 2 (7 días: {dias_7}, 8-15 días: {dias_8})...")
        
        # Procesar 7 días
        if dias_7:
            print("\n=== PROCESANDO PRÓXIMOS 7 DÍAS ===")
            if not self.selector_fecha_vencimiento_7dias():
                print("✗ Error aplicando filtro de 7 días")
                return False
            if not self._procesar_tabla("notificaciones_7_dias.xlsx"):
                return False
        
        # Procesar 8-15 días
        if dias_8:
            print("\n=== PROCESANDO 8 A 15 DÍAS ===")
            if not self.fecha_vencimiento_8dias():
                print("✗ Error aplicando filtro de 8-15 días")
                return False
            if not self._procesar_tabla("notificaciones_8_dias.xlsx"):
                return False
        
        print("\n✓ Etapa 2 completada")
        return True
    
    def _procesar_tabla(self, archivo: str) -> bool:
        """
        Procesa la tabla actual, envía notificaciones y guarda resultados en .xlsx.

        Args:
            archivo: Nombre del archivo (ej: 'notificaciones_7_dias.xlsx')

        Returns:
            bool: True si se procesó correctamente
        """
        import os
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        try:
            self.page.wait_for_selector("tr.sc-grid__table__row", timeout=15000)

            row_count = self.page.locator("tr.sc-grid__table__row").count()
            print(f"Total de clientes encontrados: {row_count}")

            resultados = []

            for i in range(row_count):
                row = self.page.locator("tr.sc-grid__table__row").nth(i)

                name_locator = row.locator("td[id^='name-and-taxid-column'] span.font-weight-bold").first
                client_name = name_locator.inner_text().strip()
                print(f"{i+1}. Cliente: {client_name}")

                pay_locator = row.locator("td[id^='payment-method-column'] span.font-weight-bold").first
                pay_text = pay_locator.inner_text().strip().upper()
                if "EFECTIVO" in pay_text:
                    pago = "EFECTIVO"
                elif "DÉBITO" in pay_text:
                    pago = "DÉBITO DIRECTO"
                elif "TARJETA" in pay_text:
                    pago = "TARJETA DE CRÉDITO"
                else:
                    pago = ""

                efectivo = "✓" if pago == "EFECTIVO" else "✗"
                debito = "✓" if pago == "DÉBITO DIRECTO" else "✗"
                tarjeta = "✓" if pago == "TARJETA DE CRÉDITO" else "✗"

                action_btn = row.locator("button.cell__dot-button").first
                action_btn.click(force=True)
                self.page.wait_for_timeout(500)
                print(f"   → Menú abierto para {client_name}")

                resultado = self._compartir()

                resultados.append({
                    'cliente': client_name,
                    'notificacion_whatsapp': '✓ enviado' if resultado['whapp'] else '✗ no enviado',
                    'notificacion_correo': '✓ enviado' if resultado['correo'] else '✗ no enviado',
                    'EFECTIVO': efectivo,
                    'DÉBITO DIRECTO': debito,
                    'TARJETA DE CRÉDITO': tarjeta,
                })

                self.page.keyboard.press("Escape")
                self.page.wait_for_timeout(500)

            report_dir = "reports"
            os.makedirs(report_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            name, _ = os.path.splitext(archivo)
            filepath = os.path.join(report_dir, f"{name}_{timestamp}.xlsx")

            wb = Workbook()
            ws = wb.active
            ws.title = "Notificaciones"

            headers = [
                'Cliente', 'WhatsApp', 'Correo',
                'EFECTIVO', 'DÉBITO DIRECTO', 'TARJETA DE CRÉDITO'
            ]
            bold_font = Font(bold=True)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = bold_font
                cell.alignment = Alignment(horizontal='center')

            for row_idx, r in enumerate(resultados, 2):
                ws.cell(row=row_idx, column=1, value=r['cliente'])
                ws.cell(row=row_idx, column=2, value=r['notificacion_whatsapp'])
                ws.cell(row=row_idx, column=3, value=r['notificacion_correo'])
                ws.cell(row=row_idx, column=4, value=r['EFECTIVO'])
                ws.cell(row=row_idx, column=5, value=r['DÉBITO DIRECTO'])
                ws.cell(row=row_idx, column=6, value=r['TARJETA DE CRÉDITO'])

            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = max_len + 4

            wb.save(filepath)
            print(f"✓ Reporte guardado: {filepath}")
            print(f"  Total procesados: {len(resultados)}")

            return True

        except Exception as e:
            print(f"Error procesando tabla: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _compartir(self) -> dict:
        """
        Compartir según configuración guardada.
        Lee 'notif_correo' y 'notif_whapp' de config.json
        """
        from config_manager import get_settings
        s = get_settings()
        correo = s.value("notif_correo", True, type=bool)
        whapp = s.value("notif_whapp", True, type=bool)
        
        result = {'correo': False, 'whapp': False}
        
        print("Iniciando proceso de compartir...")
        try:
            # Paso 1: Click en "Compartir" en el menú contextual
            print("Abriendo opción Compartir...")
            compartir_link = self.page.locator("a[id='compartir']")
            compartir_link.wait_for(state="visible", timeout=5000)
            compartir_link.click()
            print("✓ Compartir clickeado")
            self.page.wait_for_timeout(1000)
            
            # Paso 2: Flujo de Email (si correo=True)
            if correo:
                print("Flujo de Email...")
                try:
                    # Click en "Previsualizar Email"
                    preview_btn = self.page.locator("button.sc-button--secondary", has_text="Previsualizar Email")
                    preview_btn.wait_for(state="visible", timeout=5000)
                    preview_btn.click()
                    print("✓ Previsualizar Email clickeado")
                    self.page.wait_for_timeout(1500)
                    
                    # Click en "Enviar Email"
                    send_btn = self.page.locator("button.sc-button--primary", has_text="Enviar Email")
                    send_btn.wait_for(state="visible", timeout=5000)
                    send_btn.click()
                    print("✓ Enviar Email clickeado")
                    self.page.wait_for_timeout(1500)
                    
                    # Esperar mensaje de éxito
                    print("Esperando confirmación de envío...")
                    success_msg = self.page.locator("p:text('La información de pago se envió con éxito.')")
                    success_msg.wait_for(state="visible", timeout=10000)
                    print("✓ Email enviado con éxito")
                    result['correo'] = True
                    
                    self.page.wait_for_timeout(1000)
                    
                    # Click en "Cerrar" - vuelve al submenú Compartir
                    cerrar_btn = self.page.locator("button.sc-button--primary", has_text="Cerrar")
                    cerrar_btn.wait_for(state="visible", timeout=5000)
                    cerrar_btn.click()
                    print("✓ Cerrar clickeado - en submenú Compartir")
                    self.page.wait_for_timeout(1000)
                    
                except Exception as e:
                    print(f"✗ Error en flujo Email: {e}")
            
            # Paso 3: Flujo de WhatsApp (si whapp=True)
            if whapp:
                print("Flujo de WhatsApp - Capturando URL del sistema...")
                try:
                    from bot.whatsapp_manager import WhatsAppManager
                    
                    # 1. Hacer clic en "Enviar por WhatsApp" (abre api.whatsapp.com)
                    with self.page.context.expect_page(timeout=5000) as popup_info:
                        self.page.get_by_role("button", name="Enviar por WhatsApp").click()
                    
                    api_page = popup_info.value
                    api_page.wait_for_load_state("domcontentloaded")
                    print(f"✓ API WhatsApp abierto: {api_page.url[:60]}...")
                    
                    # 2. Capturar href del link "Continuar en WhatsApp Web"
                    try:
                        wa_link = api_page.get_by_role("link", name="Continuar en WhatsApp Web").first
                        wa_href = wa_link.get_attribute("href")
                        print(f"✓ URL capturada de 'Continuar en WhatsApp Web': {wa_href[:80]}...")
                        
                        # Cerrar api.whatsapp.com
                        api_page.close()
                        print("✓ Página api.whatsapp.com cerrada")
                        
                    except Exception as e:
                        print(f"✗ No se encontró el link 'Continuar en WhatsApp Web': {e}")
                        # Fallback: usar la URL de api.whatsapp.com y reemplazar
                        wa_href = api_page.url.replace("api.whatsapp.com", "web.whatsapp.com")
                        print(f"✓ Usando URL fallback: {wa_href[:80]}...")
                        api_page.close()
                    
                    # 3. Usar WhatsAppManager con la URL capturada
                    context = self.page.context
                    manager = WhatsAppManager(context)
                    
                    # Enviar mensaje usando la URL del sistema
                    if manager.send_message(url=wa_href, timeout=50000):
                        print("✓ Mensaje enviado por WhatsApp (URL del sistema)")
                        result['whapp'] = True
                    else:
                        print("✗ No se pudo enviar mensaje por WhatsApp")
                    
                    # Cerrar manager (cierra página pero NO el contexto)
                    manager.close()
                    
                except Exception as e:
                    print(f"Error en flujo WhatsApp: {e}")
                    import traceback
                    traceback.print_exc()
            
            return result
            
        except Exception as e:
            print(f"Error en _compartir: {e}")
            import traceback
            traceback.print_exc()
            return result

    def etapa_4(self):
        print("Etapa 4 - Por implementar")
        return True

    def etapa_5(self):
        print("Etapa 5 - Por implementar")
        return True

    def etapa_6(self):
        print("Etapa 6 - Por implementar")
        return True
